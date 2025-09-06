"""
Export Service

This module provides comprehensive export functionality for reports
in multiple formats including PDF, Excel, HTML, and images.
"""

import asyncio
import io
import json
import logging
import os
import tempfile
from datetime import datetime
from typing import Any, Dict, List, Optional, Union, BinaryIO
from dataclasses import asdict
import uuid

import pandas as pd
import numpy as np
from jinja2 import Environment, FileSystemLoader, Template
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.backends.backend_pdf import PdfPages
import plotly.graph_objects as go
import plotly.express as px
from plotly.utils import PlotlyJSONEncoder
import weasyprint
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from PIL import Image
import base64

from ..models.portfolio import Portfolio
from ..models.market_data import Stock
from ..utils.cache_service import CacheService

logger = logging.getLogger(__name__)


class ExportFormat:
    """Export format constants"""
    PDF = "pdf"
    EXCEL = "excel"
    HTML = "html"
    IMAGE = "image"
    CSV = "csv"
    JSON = "json"


class ExportService:
    """Comprehensive export service for reports"""
    
    def __init__(self):
        self.cache_service = CacheService()
        
        # Initialize Jinja2 environment for HTML templates
        self.jinja_env = Environment(
            loader=FileSystemLoader('templates'),
            autoescape=True,
            trim_blocks=True,
            lstrip_blocks=True
        )
        
        # Set up matplotlib style
        plt.style.use('seaborn-v0_8')
        sns.set_palette("husl")
    
    async def export_report(
        self, 
        report_content: Dict[str, Any], 
        export_format: str,
        options: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Export report in specified format"""
        try:
            logger.info(f"Starting export in format: {export_format}")
            
            if export_format == ExportFormat.PDF:
                return await self._export_to_pdf(report_content, options)
            elif export_format == ExportFormat.EXCEL:
                return await self._export_to_excel(report_content, options)
            elif export_format == ExportFormat.HTML:
                return await self._export_to_html(report_content, options)
            elif export_format == ExportFormat.IMAGE:
                return await self._export_to_image(report_content, options)
            elif export_format == ExportFormat.CSV:
                return await self._export_to_csv(report_content, options)
            elif export_format == ExportFormat.JSON:
                return await self._export_to_json(report_content, options)
            else:
                raise ValueError(f"Unsupported export format: {export_format}")
                
        except Exception as e:
            logger.error(f"Export failed: {str(e)}")
            raise
    
    async def _export_to_pdf(self, content: Dict[str, Any], options: Optional[Dict[str, Any]] = None) -> bytes:
        """Export report to PDF format"""
        try:
            # Create temporary file
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
                temp_path = temp_file.name
            
            # Generate HTML content first
            html_content = await self._generate_html_content(content, options)
            
            # Convert HTML to PDF using WeasyPrint
            weasyprint.HTML(string=html_content).write_pdf(temp_path)
            
            # Read the generated PDF
            with open(temp_path, 'rb') as f:
                pdf_content = f.read()
            
            # Clean up temporary file
            os.unlink(temp_path)
            
            return pdf_content
            
        except Exception as e:
            logger.error(f"PDF export failed: {str(e)}")
            raise
    
    async def _export_to_excel(self, content: Dict[str, Any], options: Optional[Dict[str, Any]] = None) -> bytes:
        """Export report to Excel format"""
        try:
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create summary sheet
            summary_sheet = wb.create_sheet("Summary")
            await self._populate_summary_sheet(summary_sheet, content)
            
            # Process each section
            for section in content.get("sections", []):
                if section.get("type") == "table":
                    await self._create_table_sheet(wb, section, content)
                elif section.get("type") == "chart":
                    await self._create_chart_sheet(wb, section, content)
                elif section.get("type") == "metric":
                    await self._create_metrics_sheet(wb, section, content)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Excel export failed: {str(e)}")
            raise
    
    async def _export_to_html(self, content: Dict[str, Any], options: Optional[Dict[str, Any]] = None) -> bytes:
        """Export report to HTML format"""
        try:
            # Generate HTML content
            html_content = await self._generate_html_content(content, options)
            
            return html_content.encode('utf-8')
            
        except Exception as e:
            logger.error(f"HTML export failed: {str(e)}")
            raise
    
    async def _export_to_image(self, content: Dict[str, Any], options: Optional[Dict[str, Any]] = None) -> bytes:
        """Export report to image format"""
        try:
            # Create figure
            fig, ax = plt.subplots(figsize=(12, 8))
            
            # Generate chart based on content
            await self._generate_chart_image(fig, ax, content, options)
            
            # Save to bytes
            output = io.BytesIO()
            fig.savefig(output, format='png', dpi=300, bbox_inches='tight')
            output.seek(0)
            
            plt.close(fig)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Image export failed: {str(e)}")
            raise
    
    async def _export_to_csv(self, content: Dict[str, Any], options: Optional[Dict[str, Any]] = None) -> bytes:
        """Export report to CSV format"""
        try:
            # Find table sections
            table_sections = [s for s in content.get("sections", []) if s.get("type") == "table"]
            
            if not table_sections:
                # Create a simple CSV with basic information
                df = pd.DataFrame([{
                    "Report Title": content.get("title", ""),
                    "Generated At": content.get("generated_at", ""),
                    "Type": content.get("type", "")
                }])
            else:
                # Use the first table section
                table_section = table_sections[0]
                df = await self._extract_table_data(table_section, content)
            
            # Convert to CSV
            csv_content = df.to_csv(index=False)
            
            return csv_content.encode('utf-8')
            
        except Exception as e:
            logger.error(f"CSV export failed: {str(e)}")
            raise
    
    async def _export_to_json(self, content: Dict[str, Any], options: Optional[Dict[str, Any]] = None) -> bytes:
        """Export report to JSON format"""
        try:
            # Convert content to JSON
            json_content = json.dumps(content, indent=2, default=str)
            
            return json_content.encode('utf-8')
            
        except Exception as e:
            logger.error(f"JSON export failed: {str(e)}")
            raise
    
    async def _generate_html_content(self, content: Dict[str, Any], options: Optional[Dict[str, Any]] = None) -> str:
        """Generate HTML content for the report"""
        try:
            # Load HTML template
            template = self.jinja_env.get_template('report_template.html')
            
            # Prepare context
            context = {
                "content": content,
                "options": options or {},
                "generated_at": datetime.now().isoformat(),
                "styles": self._get_report_styles()
            }
            
            # Render template
            html_content = template.render(**context)
            
            return html_content
            
        except Exception as e:
            logger.error(f"HTML generation failed: {str(e)}")
            # Fallback to simple HTML
            return self._generate_simple_html(content)
    
    def _generate_simple_html(self, content: Dict[str, Any]) -> str:
        """Generate simple HTML as fallback"""
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{content.get('title', 'Report')}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; }}
                h1 {{ color: #333; }}
                h2 {{ color: #666; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
            </style>
        </head>
        <body>
            <h1>{content.get('title', 'Report')}</h1>
            <p>Generated at: {content.get('generated_at', '')}</p>
        """
        
        for section in content.get("sections", []):
            html += f"<h2>{section.get('title', '')}</h2>"
            
            if section.get("type") == "text":
                html += f"<p>{section.get('content', '')}</p>"
            elif section.get("type") == "table":
                html += self._generate_table_html(section)
            elif section.get("type") == "metric":
                html += self._generate_metrics_html(section)
        
        html += "</body></html>"
        return html
    
    def _generate_table_html(self, section: Dict[str, Any]) -> str:
        """Generate HTML table from section"""
        # This would be implemented based on the table data structure
        return "<p>Table content would be rendered here</p>"
    
    def _generate_metrics_html(self, section: Dict[str, Any]) -> str:
        """Generate HTML metrics from section"""
        # This would be implemented based on the metrics data structure
        return "<p>Metrics content would be rendered here</p>"
    
    async def _populate_summary_sheet(self, sheet, content: Dict[str, Any]):
        """Populate Excel summary sheet"""
        try:
            # Add title
            sheet['A1'] = content.get('title', 'Report')
            sheet['A1'].font = Font(size=16, bold=True)
            
            # Add metadata
            row = 3
            sheet[f'A{row}'] = 'Generated At:'
            sheet[f'B{row}'] = content.get('generated_at', '')
            row += 1
            
            sheet[f'A{row}'] = 'Type:'
            sheet[f'B{row}'] = content.get('type', '')
            row += 1
            
            # Add section summary
            row += 1
            sheet[f'A{row}'] = 'Sections:'
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for section in content.get("sections", []):
                sheet[f'A{row}'] = section.get('title', '')
                sheet[f'B{row}'] = section.get('type', '')
                row += 1
                
        except Exception as e:
            logger.error(f"Error populating summary sheet: {str(e)}")
    
    async def _create_table_sheet(self, wb, section: Dict[str, Any], content: Dict[str, Any]):
        """Create Excel sheet for table section"""
        try:
            # Create sheet
            sheet_name = section.get('title', 'Table')[:31]  # Excel sheet name limit
            sheet = wb.create_sheet(sheet_name)
            
            # Extract table data
            df = await self._extract_table_data(section, content)
            
            if df is not None and not df.empty:
                # Add data to sheet
                for r in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(r)
                
                # Style the header
                for cell in sheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
                    
        except Exception as e:
            logger.error(f"Error creating table sheet: {str(e)}")
    
    async def _create_chart_sheet(self, wb, section: Dict[str, Any], content: Dict[str, Any]):
        """Create Excel sheet for chart section"""
        try:
            # Create sheet
            sheet_name = section.get('title', 'Chart')[:31]
            sheet = wb.create_sheet(sheet_name)
            
            # This would generate chart data and create Excel charts
            # For now, add placeholder
            sheet['A1'] = f"Chart: {section.get('title', '')}"
            sheet['A1'].font = Font(bold=True)
            
        except Exception as e:
            logger.error(f"Error creating chart sheet: {str(e)}")
    
    async def _create_metrics_sheet(self, wb, section: Dict[str, Any], content: Dict[str, Any]):
        """Create Excel sheet for metrics section"""
        try:
            # Create sheet
            sheet_name = section.get('title', 'Metrics')[:31]
            sheet = wb.create_sheet(sheet_name)
            
            # Add metrics data
            metrics = section.get('metrics', [])
            
            sheet['A1'] = 'Metric'
            sheet['B1'] = 'Value'
            sheet['A1'].font = Font(bold=True)
            sheet['B1'].font = Font(bold=True)
            
            row = 2
            for metric in metrics:
                sheet[f'A{row}'] = metric.get('name', '')
                sheet[f'B{row}'] = metric.get('value', '')
                row += 1
                
        except Exception as e:
            logger.error(f"Error creating metrics sheet: {str(e)}")
    
    async def _extract_table_data(self, section: Dict[str, Any], content: Dict[str, Any]) -> Optional[pd.DataFrame]:
        """Extract table data from section"""
        try:
            data_source = section.get('data_source')
            if not data_source:
                return None
            
            # This would extract data based on the data source
            # For now, return a sample DataFrame
            return pd.DataFrame({
                'Column 1': [1, 2, 3],
                'Column 2': ['A', 'B', 'C'],
                'Column 3': [10.5, 20.3, 30.1]
            })
            
        except Exception as e:
            logger.error(f"Error extracting table data: {str(e)}")
            return None
    
    async def _generate_chart_image(self, fig, ax, content: Dict[str, Any], options: Optional[Dict[str, Any]] = None):
        """Generate chart image"""
        try:
            # Find chart sections
            chart_sections = [s for s in content.get("sections", []) if s.get("type") == "chart"]
            
            if chart_sections:
                chart_section = chart_sections[0]
                await self._create_chart_from_section(fig, ax, chart_section, content)
            else:
                # Create a simple summary chart
                ax.text(0.5, 0.5, f"Report: {content.get('title', '')}", 
                       ha='center', va='center', fontsize=16, transform=ax.transAxes)
                ax.set_title("Report Summary")
            
        except Exception as e:
            logger.error(f"Error generating chart image: {str(e)}")
    
    async def _create_chart_from_section(self, fig, ax, section: Dict[str, Any], content: Dict[str, Any]):
        """Create chart from section data"""
        try:
            chart_type = section.get('chart_type', 'line')
            data_source = section.get('data_source')
            
            # This would create actual charts based on the data
            # For now, create a sample chart
            if chart_type == 'line':
                x = [1, 2, 3, 4, 5]
                y = [2, 4, 1, 5, 3]
                ax.plot(x, y, marker='o')
            elif chart_type == 'bar':
                x = ['A', 'B', 'C', 'D']
                y = [3, 7, 2, 5]
                ax.bar(x, y)
            elif chart_type == 'pie':
                labels = ['A', 'B', 'C', 'D']
                sizes = [30, 25, 20, 25]
                ax.pie(sizes, labels=labels, autopct='%1.1f%%')
            
            ax.set_title(section.get('title', 'Chart'))
            
        except Exception as e:
            logger.error(f"Error creating chart from section: {str(e)}")
    
    def _get_report_styles(self) -> str:
        """Get CSS styles for HTML reports"""
        return """
        <style>
            body {
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                line-height: 1.6;
                color: #333;
                max-width: 1200px;
                margin: 0 auto;
                padding: 20px;
            }
            .header {
                border-bottom: 2px solid #007acc;
                padding-bottom: 20px;
                margin-bottom: 30px;
            }
            .header h1 {
                color: #007acc;
                margin: 0;
                font-size: 2.5em;
            }
            .header .meta {
                color: #666;
                font-size: 0.9em;
                margin-top: 10px;
            }
            .section {
                margin-bottom: 40px;
                page-break-inside: avoid;
            }
            .section h2 {
                color: #007acc;
                border-bottom: 1px solid #ddd;
                padding-bottom: 10px;
                margin-bottom: 20px;
            }
            .metrics-grid {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 20px;
                margin-bottom: 30px;
            }
            .metric-card {
                background: #f8f9fa;
                border: 1px solid #e9ecef;
                border-radius: 8px;
                padding: 20px;
                text-align: center;
            }
            .metric-value {
                font-size: 2em;
                font-weight: bold;
                color: #007acc;
                margin-bottom: 5px;
            }
            .metric-label {
                color: #666;
                font-size: 0.9em;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 20px;
            }
            th, td {
                border: 1px solid #ddd;
                padding: 12px;
                text-align: left;
            }
            th {
                background-color: #f8f9fa;
                font-weight: bold;
                color: #333;
            }
            tr:nth-child(even) {
                background-color: #f8f9fa;
            }
            .chart-container {
                margin: 20px 0;
                text-align: center;
            }
            .footer {
                margin-top: 50px;
                padding-top: 20px;
                border-top: 1px solid #ddd;
                color: #666;
                font-size: 0.9em;
                text-align: center;
            }
            @media print {
                body { margin: 0; }
                .section { page-break-inside: avoid; }
            }
        </style>
        """
    
    async def batch_export(
        self, 
        reports: List[Dict[str, Any]], 
        export_format: str,
        options: Optional[Dict[str, Any]] = None
    ) -> Dict[str, bytes]:
        """Export multiple reports in batch"""
        try:
            results = {}
            
            for i, report in enumerate(reports):
                try:
                    export_data = await self.export_report(report, export_format, options)
                    results[f"report_{i+1}"] = export_data
                except Exception as e:
                    logger.error(f"Failed to export report {i+1}: {str(e)}")
                    results[f"report_{i+1}"] = None
            
            return results
            
        except Exception as e:
            logger.error(f"Batch export failed: {str(e)}")
            raise
    
    async def get_export_info(self, export_format: str) -> Dict[str, Any]:
        """Get information about export format capabilities"""
        info = {
            "format": export_format,
            "supported": True,
            "mime_type": self._get_mime_type(export_format),
            "file_extension": self._get_file_extension(export_format),
            "capabilities": self._get_format_capabilities(export_format)
        }
        
        return info
    
    def _get_mime_type(self, export_format: str) -> str:
        """Get MIME type for export format"""
        mime_types = {
            ExportFormat.PDF: "application/pdf",
            ExportFormat.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ExportFormat.HTML: "text/html",
            ExportFormat.IMAGE: "image/png",
            ExportFormat.CSV: "text/csv",
            ExportFormat.JSON: "application/json"
        }
        return mime_types.get(export_format, "application/octet-stream")
    
    def _get_file_extension(self, export_format: str) -> str:
        """Get file extension for export format"""
        extensions = {
            ExportFormat.PDF: ".pdf",
            ExportFormat.EXCEL: ".xlsx",
            ExportFormat.HTML: ".html",
            ExportFormat.IMAGE: ".png",
            ExportFormat.CSV: ".csv",
            ExportFormat.JSON: ".json"
        }
        return extensions.get(export_format, ".bin")
    
    def _get_format_capabilities(self, export_format: str) -> List[str]:
        """Get capabilities for export format"""
        capabilities = {
            ExportFormat.PDF: ["text", "tables", "charts", "images", "styling"],
            ExportFormat.EXCEL: ["tables", "charts", "formulas", "multiple_sheets"],
            ExportFormat.HTML: ["text", "tables", "charts", "interactive", "responsive"],
            ExportFormat.IMAGE: ["charts", "visualizations"],
            ExportFormat.CSV: ["tables", "data"],
            ExportFormat.JSON: ["structured_data", "metadata"]
        }
        return capabilities.get(export_format, [])
